import json, io, csv
from datetime import datetime, timedelta
from collections import defaultdict
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils import timezone
from django.contrib.auth import login, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import (
    LoginView, LogoutView, PasswordResetView, PasswordResetDoneView,
    PasswordResetConfirmView, PasswordResetCompleteView,
)
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.db.models import Q, Count, Avg, F, DurationField, ExpressionWrapper
from django.db.models.functions import TruncMonth
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from .models import User, LoanApplication, DocumentUpload, ApplicationHistory, AuditLog, INDIAN_STATES
from .forms import (
    LoginForm, UserForm, CustomPasswordResetForm, CustomSetPasswordForm,
    LoanApplicationForm, DocumentUploadForm, ApplicationRemarkForm,
)
from .decorators import role_required, super_admin_required, gcc_or_admin_required

ROLE_DASHBOARD = {
    'SUPER_ADMIN': reverse_lazy('dashboard'),
    'GCC_NOIDA': reverse_lazy('gcc_dashboard'),
    'STATE_HEAD': reverse_lazy('state_head_dashboard'),
    'COORDINATOR': reverse_lazy('coordinator_dashboard'),
}


class CustomLoginView(LoginView):
    form_class = LoginForm
    template_name = 'registration/login.html'
    redirect_authenticated_user = True

    def form_valid(self, form):
        response = super().form_valid(form)
        AuditLog.objects.create(
            user=self.request.user, action='LOGIN', module='Auth',
            ip_address=self.get_client_ip(),
        )
        return response

    def get_client_ip(self):
        x_fwd = self.request.META.get('HTTP_X_FORWARDED_FOR')
        if x_fwd:
            return x_fwd.split(',')[0].strip()
        return self.request.META.get('REMOTE_ADDR')

    def get_success_url(self):
        if self.request.user.is_superuser:
            return str(reverse_lazy('dashboard'))
        return str(ROLE_DASHBOARD.get(self.request.user.role, reverse_lazy('dashboard')))


class CustomLogoutView(LogoutView):
    http_method_names = ['get', 'post']

    def get(self, request, *args, **kwargs):
        from django.contrib.auth import logout
        logout(request)
        if request.GET.get('timeout'):
            return render(request, 'registration/session_expired.html')
        next_page = self.get_next_page()
        return redirect(next_page) if next_page else redirect('login')


class CustomPasswordResetView(PasswordResetView):
    form_class = CustomPasswordResetForm
    template_name = 'registration/password_reset_form.html'
    email_template_name = 'registration/password_reset_email.txt'
    subject_template_name = 'registration/password_reset_subject.txt'
    success_url = reverse_lazy('password_reset_done')


class CustomPasswordResetDoneView(PasswordResetDoneView):
    template_name = 'registration/password_reset_done.html'


class CustomPasswordResetConfirmView(PasswordResetConfirmView):
    form_class = CustomSetPasswordForm
    template_name = 'registration/password_reset_confirm.html'
    success_url = reverse_lazy('password_reset_complete')


class CustomPasswordResetCompleteView(PasswordResetCompleteView):
    template_name = 'registration/password_reset_complete.html'


@login_required
def dashboard(request):
    user = request.user
    if user.is_superuser or user.role == 'SUPER_ADMIN':
        applications = LoanApplication.objects.all()
        now = timezone.now()
        seven_days_ago = now - timedelta(days=7)
        fourteen_days_ago = now - timedelta(days=14)
        stalled = (
            Q(status__in=['SUBMITTED', 'RESUBMITTED', 'GCC_REVIEW', 'ADDITIONAL_DOCS_REQUIRED'])
            & Q(updated_at__lte=seven_days_ago)
        )
        stalled_apps = applications.filter(stalled)
        aged_apps = []
        for app in applications.filter(status__in=['SUBMITTED', 'RESUBMITTED', 'GCC_REVIEW', 'ADDITIONAL_DOCS_REQUIRED', 'RETURNED_FOR_CORRECTION']):
            days = (now - app.updated_at).days
            if days >= 14:
                level = 'critical'
            elif days >= 7:
                level = 'warning'
            else:
                level = 'normal'
            aged_apps.append({'app': app, 'days': days, 'level': level})
        aged_apps.sort(key=lambda x: x['days'], reverse=True)
        activity_timeline = ApplicationHistory.objects.select_related('application', 'changed_by').order_by('-created_at')[:15]
        context = {
            'total_applications': applications.count(),
            'pending_review': applications.filter(status__in=['SUBMITTED', 'RESUBMITTED']).count(),
            'gcc_review': applications.filter(status='GCC_REVIEW').count(),
            'additional_docs_req': applications.filter(status='ADDITIONAL_DOCS_REQUIRED').count(),
            'returned': applications.filter(status='RETURNED_FOR_CORRECTION').count(),
            'approved': applications.filter(status='APPROVED').count(),
            'rejected': applications.filter(status='REJECTED').count(),
            'disbursed': applications.filter(status='DISBURSED').count(),
            'stalled_count': stalled_apps.count(),
            'aged_apps': aged_apps[:10],
            'activity_timeline': activity_timeline,
            'recent_applications': applications.order_by('-created_at')[:10],
            'role_counts': {
                role: User.objects.filter(role=role).count()
                for role, _ in User.Role.choices
            },
        }
        return render(request, 'dashboard.html', context)
    if user.role == 'STATE_HEAD':
        return state_head_dashboard(request)
    if user.role == 'GCC_NOIDA':
        return gcc_dashboard(request)
    if user.role == 'COORDINATOR':
        applications = LoanApplication.objects.filter(created_by=user)
        app_id_filter = request.GET.get('application_id', '')
        coord_status_filter = request.GET.get('status', '')
        coord_date_from = request.GET.get('date_from', '')
        coord_date_to = request.GET.get('date_to', '')
        if app_id_filter:
            applications = applications.filter(application_id__icontains=app_id_filter)
        if coord_status_filter:
            applications = applications.filter(status=coord_status_filter)
        if coord_date_from:
            applications = applications.filter(created_at__gte=coord_date_from)
        if coord_date_to:
            applications = applications.filter(created_at__lte=coord_date_to + ' 23:59:59')
        total = applications.count()
        coord_page = request.GET.get('page', 1)
        coord_paginator = Paginator(applications.order_by('-created_at'), 15)
        try:
            coord_page_obj = coord_paginator.page(coord_page)
        except (PageNotAnInteger, EmptyPage):
            coord_page_obj = coord_paginator.page(1)
        now = timezone.now()
        aged_apps = []
        for app in applications.filter(status__in=['SUBMITTED', 'RESUBMITTED', 'RETURNED_FOR_CORRECTION', 'ADDITIONAL_DOCS_REQUIRED']):
            days = (now - app.updated_at).days
            if days >= 14: level = 'critical'
            elif days >= 7: level = 'warning'
            else: level = 'normal'
            aged_apps.append({'app': app, 'days': days, 'level': level})
        stalled_count = sum(1 for a in aged_apps if a['level'] != 'normal')
        activity_timeline = ApplicationHistory.objects.filter(application__in=applications).select_related('application', 'changed_by').order_by('-created_at')[:10]
        context = {
            'total_applications': total,
            'draft_count': applications.filter(status='DRAFT').count(),
            'pending_review': applications.filter(status__in=['SUBMITTED', 'RESUBMITTED']).count(),
            'returned': applications.filter(status='RETURNED_FOR_CORRECTION').count(),
            'additional_docs_required': applications.filter(status='ADDITIONAL_DOCS_REQUIRED').count(),
            'approved': applications.filter(status='APPROVED').count(),
            'rejected': applications.filter(status='REJECTED').count(),
            'disbursed': applications.filter(status='DISBURSED').count(),
            'stalled_count': stalled_count,
            'aged_apps': aged_apps[:8],
            'activity_timeline': activity_timeline,
            'coord_page_obj': coord_page_obj,
            'status_choices': LoanApplication.Status.choices,
            'app_id_filter': app_id_filter, 'coord_status_filter': coord_status_filter,
            'coord_date_from': coord_date_from, 'coord_date_to': coord_date_to,
            'filter_params': request.GET.urlencode() if request.GET else '',
        }
        return render(request, 'coordinator/dashboard.html', context)
    context = {
        'total_applications': applications.count(),
        'under_review': applications.filter(status__in=['SUBMITTED', 'STATE_REVIEW', 'RESUBMITTED', 'GCC_REVIEW', 'ADDITIONAL_DOCS_REQUIRED']).count(),
        'returned': applications.filter(status='RETURNED_FOR_CORRECTION').count(),
        'approved': applications.filter(status='APPROVED').count(),
        'rejected': applications.filter(status='REJECTED').count(),
        'recent_applications': applications.order_by('-created_at')[:10],
    }
    return render(request, 'dashboard.html', context)


# ─── Loan List (Super Admin / GCC) ──────────────────────────────

@login_required
def loan_list(request):
    if not (request.user.is_superuser or request.user.role in ('SUPER_ADMIN', 'GCC_NOIDA', 'STATE_HEAD')):
        if request.user.role == 'COORDINATOR':
            return redirect('coordinator_dashboard')
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    base_qs = LoanApplication.objects.all().select_related('created_by')
    if request.user.role == 'STATE_HEAD' and request.user.state:
        base_qs = base_qs.filter(state=request.user.state)
    search = request.GET.get('q', '')
    state_filter = request.GET.get('state', '')
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if search:
        base_qs = base_qs.filter(
            Q(application_id__icontains=search) |
            Q(applicant_name__icontains=search) |
            Q(created_by__first_name__icontains=search) |
            Q(created_by__last_name__icontains=search)
        )
    if state_filter:
        base_qs = base_qs.filter(state=state_filter)
    if status_filter:
        base_qs = base_qs.filter(status=status_filter)
    if date_from:
        base_qs = base_qs.filter(created_at__gte=date_from)
    if date_to:
        base_qs = base_qs.filter(created_at__lte=date_to + ' 23:59:59')
    all_qs = LoanApplication.objects.all()
    if request.user.role == 'STATE_HEAD' and request.user.state:
        all_qs = all_qs.filter(state=request.user.state)
    state_summary = (
        all_qs.values('state')
        .annotate(
            total=Count('id'),
            submitted=Count('id', filter=Q(status='SUBMITTED')),
            state_review=Count('id', filter=Q(status='STATE_REVIEW')),
            returned=Count('id', filter=Q(status='RETURNED_FOR_CORRECTION')),
            resubmitted=Count('id', filter=Q(status='RESUBMITTED')),
            gcc_review=Count('id', filter=Q(status='GCC_REVIEW')),
            additional_docs=Count('id', filter=Q(status='ADDITIONAL_DOCS_REQUIRED')),
            approved=Count('id', filter=Q(status='APPROVED')),
            rejected=Count('id', filter=Q(status='REJECTED')),
            disbursed=Count('id', filter=Q(status='DISBURSED')),
        )
        .order_by('-total')
    )
    total_approved = all_qs.filter(status='APPROVED').count()
    total_rejected = all_qs.filter(status='REJECTED').count()
    total_disbursed = all_qs.filter(status='DISBURSED').count()
    total_pending = all_qs.filter(status__in=['SUBMITTED', 'STATE_REVIEW', 'RESUBMITTED', 'GCC_REVIEW', 'ADDITIONAL_DOCS_REQUIRED']).count()
    total_count = all_qs.count()
    ll_page = request.GET.get('page', 1)
    ll_paginator = Paginator(base_qs.order_by('-created_at'), 20)
    try:
        ll_page_obj = ll_paginator.page(ll_page)
    except (PageNotAnInteger, EmptyPage):
        ll_page_obj = ll_paginator.page(1)
    get_params = request.GET.copy()
    if 'page' in get_params: del get_params['page']
    context = {
        'll_page_obj': ll_page_obj,
        'filter_params': get_params.urlencode() if get_params else '',
        'states_list': [s[0] for s in INDIAN_STATES],
        'status_choices': LoanApplication.Status.choices,
        'search_query': search, 'state_filter': state_filter,
        'status_filter': status_filter, 'date_from': date_from, 'date_to': date_to,
        'state_summary': state_summary,
        'total_count': total_count,
        'total_pending': total_pending,
        'total_approved': total_approved,
        'total_rejected': total_rejected,
        'total_disbursed': total_disbursed,
        'state_summary_json': json.dumps([{
            'state': s['state'] or 'Unknown',
            'total': s['total'],
            'approved': s['approved'],
            'rejected': s['rejected'],
            'pending': s['submitted'] + s['state_review'] + s['resubmitted'] + s['gcc_review'] + s['additional_docs'],
            'returned': s['returned'],
            'disbursed': s['disbursed'],
        } for s in state_summary]),
    }
    return render(request, 'loan_list.html', context)


# ─── State Head ─────────────────────────────────────────────────

@login_required
@role_required('STATE_HEAD')
def state_head_dashboard(request):
    user = request.user
    if not user.state:
        messages.warning(request, 'Your profile has no state assigned. Please contact a Super Admin to set your state.')
        return render(request, 'state_head/dashboard.html', {
            'user_state_missing': True,
            'total_applications': 0, 'pending_review': 0, 'returned_correction': 0,
            'forwarded_gcc': 0, 'additional_docs_req': 0, 'approved': 0,
            'rejected': 0, 'disbursed': 0,
            'applications': [], 'coordinators': [], 'districts': [],
            'status_choices': LoanApplication.Status.choices,
            'status_counts_json': '{}', 'district_data_json': '[]',
        })
    base_qs = LoanApplication.objects.filter(state=user.state)
    district_filter = request.GET.get('district', '')
    coordinator_filter = request.GET.get('coordinator', '')
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    applications = base_qs
    if district_filter:
        applications = applications.filter(district=district_filter)
    if coordinator_filter:
        applications = applications.filter(created_by_id=coordinator_filter)
    if status_filter:
        applications = applications.filter(status=status_filter)
    if date_from:
        applications = applications.filter(created_at__gte=date_from)
    if date_to:
        applications = applications.filter(created_at__lte=date_to + ' 23:59:59')
    coordinators = User.objects.filter(role='COORDINATOR', state=user.state, is_active=True).order_by('first_name')
    districts = base_qs.values_list('district', flat=True).distinct().order_by('district')
    status_counts = {s[0]: base_qs.filter(status=s[0]).count() for s in LoanApplication.Status.choices}
    district_data = base_qs.values('district').annotate(count=Count('id')).order_by('-count')
    sh_page = request.GET.get('page', 1)
    sh_paginator = Paginator(applications.order_by('-created_at'), 15)
    try:
        sh_page_obj = sh_paginator.page(sh_page)
    except (PageNotAnInteger, EmptyPage):
        sh_page_obj = sh_paginator.page(1)
    now = timezone.now()
    aged_apps = []
    for app in base_qs.filter(status__in=['SUBMITTED', 'RESUBMITTED', 'RETURNED_FOR_CORRECTION', 'ADDITIONAL_DOCS_REQUIRED']):
        days = (now - app.updated_at).days
        if days >= 14: level = 'critical'
        elif days >= 7: level = 'warning'
        else: level = 'normal'
        aged_apps.append({'app': app, 'days': days, 'level': level})
    stalled_count = sum(1 for a in aged_apps if a['level'] != 'normal')
    activity_timeline = ApplicationHistory.objects.filter(application__in=base_qs).select_related('application', 'changed_by').order_by('-created_at')[:10]
    context = {
        'total_applications': base_qs.count(),
        'pending_review': base_qs.filter(status__in=['SUBMITTED', 'RESUBMITTED']).count(),
        'returned_correction': base_qs.filter(status='RETURNED_FOR_CORRECTION').count(),
        'forwarded_gcc': base_qs.filter(status='GCC_REVIEW').count(),
        'additional_docs_req': base_qs.filter(status='ADDITIONAL_DOCS_REQUIRED').count(),
        'approved': base_qs.filter(status='APPROVED').count(),
        'rejected': base_qs.filter(status='REJECTED').count(),
        'disbursed': base_qs.filter(status='DISBURSED').count(),
        'stalled_count': stalled_count,
        'aged_apps': aged_apps[:8],
        'activity_timeline': activity_timeline,
        'sh_page_obj': sh_page_obj,
        'applications': applications.order_by('-created_at'),
        'coordinators': coordinators, 'districts': districts,
        'status_choices': LoanApplication.Status.choices,
        'district_filter': district_filter, 'coordinator_filter': coordinator_filter,
        'status_filter': status_filter, 'date_from': date_from, 'date_to': date_to,
        'status_counts_json': json.dumps(status_counts),
        'district_data_json': json.dumps([{'district': d['district'], 'count': d['count']} for d in district_data]),
        'filter_params': request.GET.urlencode() if request.GET else '',
    }
    return render(request, 'state_head/dashboard.html', context)


@login_required
@role_required('STATE_HEAD')
def state_head_review(request, application_id):
    application = get_object_or_404(LoanApplication, id=application_id, state=request.user.state)
    if application.status not in ('SUBMITTED', 'RESUBMITTED'):
        messages.error(request, 'This application is not available for review.')
        return redirect('state_head_dashboard')
    documents = application.documents.all()
    history = application.history.all()
    seen = set()
    status_path = []
    for h in history:
        if h.to_status and h.to_status not in seen:
            status_path.append(h.to_status)
            seen.add(h.to_status)
        if h.from_status and h.from_status not in seen:
            status_path.append(h.from_status)
            seen.add(h.from_status)
    if application.status not in seen:
        status_path.append(application.status)
    status_labels = dict(LoanApplication.Status.choices)
    if request.method == 'POST':
        action = request.POST.get('action')
        remarks = request.POST.get('remarks', '').strip()
        if not remarks:
            messages.error(request, 'Remarks are required for all review actions.')
            return redirect('state_head_review', application_id=application.id)
        transitions = {
            'approve': ('APPROVED', 'Application approved by State Head.'),
            'forward_gcc': ('GCC_REVIEW', 'Forwarded to GCC Noida for review.'),
            'return_correction': ('RETURNED_FOR_CORRECTION', 'Returned for correction by State Head.'),
            'request_docs': ('ADDITIONAL_DOCS_REQUIRED', 'Additional documents requested by State Head.'),
        }
        if action in transitions:
            from_status = application.status
            to_status, default_remark = transitions[action]
            ApplicationHistory.objects.create(
                application=application, from_status=from_status, to_status=to_status,
                remarks=remarks or default_remark, changed_by=request.user,
            )
            application.status = to_status
            application.save()
            action_labels = {'approve': 'approved', 'forward_gcc': 'forwarded to GCC',
                             'return_correction': 'returned for correction', 'request_docs': 'requested additional documents for'}
            messages.success(request, f'Application {action_labels.get(action, "updated")} successfully.')
            return redirect('state_head_dashboard')
    return render(request, 'state_head/review.html', {
        'application': application, 'documents': documents, 'history': history,
        'status_path': status_path, 'status_labels': status_labels,
        'activity_log': get_activity_log(application),
    })


# ─── GCC Noida ──────────────────────────────────────────────────

@login_required
@role_required('GCC_NOIDA', 'SUPER_ADMIN')
def gcc_dashboard(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        app_id = request.POST.get('application_id')
        remarks = request.POST.get('remarks', '').strip()
        if action and app_id and remarks:
            application = get_object_or_404(LoanApplication, id=app_id)
            transitions = {
                'quick_approve': ('GCC_REVIEW', 'APPROVED', 'Approved via dashboard.'),
                'quick_reject': ('GCC_REVIEW', 'REJECTED', 'Rejected via dashboard.'),
                'quick_disburse': ('APPROVED', 'DISBURSED', 'Marked as disbursed via dashboard.'),
                'quick_docs': ('GCC_REVIEW', 'ADDITIONAL_DOCS_REQUIRED', 'Additional documents requested.'),
            }
            if action in transitions:
                expected_from, to_status, default_remark = transitions[action]
                if application.status == expected_from:
                    ApplicationHistory.objects.create(
                        application=application, from_status=application.status,
                        to_status=to_status, remarks=remarks or default_remark,
                        changed_by=request.user,
                    )
                    application.status = to_status
                    application.save()
                    messages.success(request, f'{application.application_id} updated successfully.')
                else:
                    messages.error(request, f'Cannot {action} — application is in {application.get_status_display()}.')
        return redirect('gcc_dashboard')

    base_qs = LoanApplication.objects.all()
    search_query = request.GET.get('q', '')
    state_filter = request.GET.get('state', '')
    status_filter = request.GET.get('status', '')
    loan_type_filter = request.GET.get('loan_type', '')
    show_all = request.GET.get('all', '')

    applications = base_qs
    if not show_all:
        applications = applications.filter(status__in=['GCC_REVIEW', 'ADDITIONAL_DOCS_REQUIRED', 'APPROVED'])
    if search_query:
        applications = applications.filter(
            Q(application_id__icontains=search_query) |
            Q(applicant_name__icontains=search_query) |
            Q(created_by__first_name__icontains=search_query) |
            Q(created_by__last_name__icontains=search_query)
        )
    if state_filter:
        applications = applications.filter(state=state_filter)
    if status_filter:
        applications = applications.filter(status=status_filter)
    if loan_type_filter:
        applications = applications.filter(loan_type=loan_type_filter)

    status_counts = {s[0]: base_qs.filter(status=s[0]).count() for s in LoanApplication.Status.choices}

    state_data = base_qs.values('state').annotate(count=Count('id')).order_by('-count')

    monthly = (
        base_qs.annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )

    approved_qs = base_qs.filter(status='APPROVED', submitted_at__isnull=False).exclude(submitted_at__isnull=True)
    turnaround_days = None
    turnaround_labels = []
    turnaround_values = []
    if approved_qs.exists():
        monthly_approved = (
            approved_qs.annotate(month=TruncMonth('submitted_at'))
            .values('month')
            .annotate(
                avg_days=Avg(
                    ExpressionWrapper(
                        F('updated_at') - F('submitted_at'),
                        output_field=DurationField()
                    )
                )
            )
            .order_by('month')
        )
        for entry in monthly_approved:
            if entry['avg_days']:
                turnaround_labels.append(entry['month'].strftime('%b %Y') if entry['month'] else '')
                turnaround_values.append(round(entry['avg_days'].days, 1))

    gcc_page = request.GET.get('page', 1)
    gcc_paginator = Paginator(applications.order_by('-created_at'), 15)
    try:
        gcc_page_obj = gcc_paginator.page(gcc_page)
    except (PageNotAnInteger, EmptyPage):
        gcc_page_obj = gcc_paginator.page(1)

    now = timezone.now()
    aged_apps = []
    for app in base_qs.filter(status__in=['GCC_REVIEW', 'ADDITIONAL_DOCS_REQUIRED', 'APPROVED', 'SUBMITTED', 'RESUBMITTED']):
        days = (now - app.updated_at).days
        if days >= 14: level = 'critical'
        elif days >= 7: level = 'warning'
        else: level = 'normal'
        aged_apps.append({'app': app, 'days': days, 'level': level})
    stalled_count = sum(1 for a in aged_apps if a['level'] != 'normal')
    activity_timeline = ApplicationHistory.objects.select_related('application', 'changed_by').order_by('-created_at')[:10]

    pending_review_count = base_qs.filter(status='GCC_REVIEW').count()
    pending_docs_count = base_qs.filter(status='ADDITIONAL_DOCS_REQUIRED').count()
    pending_disburse_count = base_qs.filter(status='APPROVED').count()
    context = {
        'total_applications': base_qs.count(),
        'pending_gcc_review': pending_review_count + pending_docs_count,
        'pending_review': pending_review_count,
        'pending_docs': pending_docs_count,
        'pending_disburse': pending_disburse_count,
        'total_pending_actions': pending_review_count + pending_docs_count + pending_disburse_count,
        'approved': base_qs.filter(status='APPROVED').count(),
        'rejected': base_qs.filter(status='REJECTED').count(),
        'disbursed': base_qs.filter(status='DISBURSED').count(),
        'stalled_count': stalled_count,
        'aged_apps': aged_apps[:8],
        'activity_timeline': activity_timeline,
        'gcc_page_obj': gcc_page_obj,
        'applications': applications.order_by('-created_at'),
        'states_list': [s[0] for s in INDIAN_STATES],
        'status_choices': LoanApplication.Status.choices,
        'loan_type_choices': LoanApplication.LoanType.choices,
        'state_filter': state_filter, 'status_filter': status_filter,
        'loan_type_filter': loan_type_filter, 'search_query': search_query, 'show_all': show_all,
        'status_counts_json': json.dumps(status_counts),
        'state_data_json': json.dumps([{'state': d['state'], 'count': d['count']} for d in state_data]),
        'monthly_labels_json': json.dumps([m['month'].strftime('%b %Y') if m['month'] else '' for m in monthly]),
        'monthly_values_json': json.dumps([m['count'] for m in monthly]),
        'turnaround_labels_json': json.dumps(turnaround_labels),
        'turnaround_values_json': json.dumps(turnaround_values),
    }
    return render(request, 'gcc/dashboard.html', context)


@login_required
@role_required('GCC_NOIDA', 'SUPER_ADMIN')
def gcc_review(request, application_id):
    application = get_object_or_404(LoanApplication, id=application_id)
    documents = application.documents.all()
    history = application.history.all()

    seen = set()
    status_path = []
    for h in history:
        if h.to_status and h.to_status not in seen:
            status_path.append(h.to_status)
            seen.add(h.to_status)
        if h.from_status and h.from_status not in seen:
            status_path.append(h.from_status)
            seen.add(h.from_status)
    if application.status not in seen:
        status_path.append(application.status)
    status_labels = dict(LoanApplication.Status.choices)

    if request.method == 'POST':
        action = request.POST.get('action')
        remarks = request.POST.get('remarks', '').strip()
        if not remarks:
            messages.error(request, 'Remarks are required for all review actions.')
            return redirect('gcc_review', application_id=application.id)

        transitions = {
            'approve': ('GCC_REVIEW', 'APPROVED', 'Application approved by GCC Noida.'),
            'reject': ('GCC_REVIEW', 'REJECTED', 'Application rejected by GCC Noida.'),
            'request_docs': ('ADDITIONAL_DOCS_REQUIRED', 'Additional documents requested by GCC Noida.'),
            'disburse': ('APPROVED', 'DISBURSED', 'Loan disbursed by GCC Noida.'),
        }

        if action in transitions:
            from_status = application.status
            expected_from, to_status, default_remark = transitions[action]
            if from_status != expected_from:
                messages.error(request, f'Cannot {action} application in current status.')
                return redirect('gcc_review', application_id=application.id)
            ApplicationHistory.objects.create(
                application=application, from_status=from_status, to_status=to_status,
                remarks=remarks or default_remark, changed_by=request.user,
            )
            application.status = to_status
            application.save()
            action_labels = {'approve': 'approved', 'reject': 'rejected',
                             'request_docs': 'requested additional documents for', 'disburse': 'disbursed'}
            messages.success(request, f'Application {action_labels.get(action, "updated")} successfully.')
            return redirect('gcc_dashboard')

    return render(request, 'gcc/review.html', {
        'application': application, 'documents': documents, 'history': history,
        'status_path': status_path, 'status_labels': status_labels,
        'activity_log': get_activity_log(application),
    })


@login_required
@require_POST
def document_verify(request, application_id, doc_id):
    application = get_object_or_404(LoanApplication, id=application_id)
    doc = get_object_or_404(DocumentUpload, id=doc_id, application=application)
    if not (request.user.role in ('GCC_NOIDA', 'STATE_HEAD', 'SUPER_ADMIN') or request.user.is_superuser):
        messages.error(request, 'Permission denied.')
        return redirect('loan_detail', application_id=application.id)
    if request.user.role == 'STATE_HEAD' and application.state != request.user.state:
        messages.error(request, 'You can only verify documents for your state.')
        return redirect('state_head_dashboard')
    status = request.POST.get('status')
    remarks = request.POST.get('remarks', '').strip()
    if status in ('VERIFIED', 'REJECTED'):
        doc.verification_status = status
        doc.verification_remarks = remarks or ('Verified' if status == 'VERIFIED' else 'Rejected')
        doc.verified_by = request.user
        doc.verified_at = timezone.now()
        doc.save()
        AuditLog.objects.create(
            user=request.user, action='VERIFY_DOCUMENT', module='Document Verification',
            remarks=f'{status} document {doc.get_doc_type_display()} for {application.application_id}',
        )
        messages.success(request, f'Document {status.lower()} successfully.')
    else:
        messages.error(request, 'Invalid verification status.')
    return redirect(request.META.get('HTTP_REFERER', f'/gcc/{application.id}/review/'))


# ─── Bulk Actions ────────────────────────────────────────────────

@login_required
@require_POST
def bulk_action(request):
    action = request.POST.get('action')
    ids = request.POST.getlist('selected_ids')
    remarks = request.POST.get('remarks', '').strip()
    if not ids or not action:
        messages.error(request, 'No applications selected or no action specified.')
        return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
    apps = LoanApplication.objects.filter(id__in=ids)
    count = 0
    if action == 'bulk_approve' and (request.user.role in ('GCC_NOIDA', 'SUPER_ADMIN') or request.user.is_superuser):
        for app in apps.filter(status='GCC_REVIEW'):
            ApplicationHistory.objects.create(
                application=app, from_status=app.status, to_status='APPROVED',
                remarks=remarks or 'Bulk approved.', changed_by=request.user,
            )
            app.status = 'APPROVED'; app.save(); count += 1
    elif action == 'bulk_reject' and (request.user.role in ('GCC_NOIDA', 'SUPER_ADMIN') or request.user.is_superuser):
        for app in apps.filter(status='GCC_REVIEW'):
            ApplicationHistory.objects.create(
                application=app, from_status=app.status, to_status='REJECTED',
                remarks=remarks or 'Bulk rejected.', changed_by=request.user,
            )
            app.status = 'REJECTED'; app.save(); count += 1
    elif action == 'bulk_disburse' and (request.user.role in ('GCC_NOIDA', 'SUPER_ADMIN') or request.user.is_superuser):
        for app in apps.filter(status='APPROVED'):
            ApplicationHistory.objects.create(
                application=app, from_status=app.status, to_status='DISBURSED',
                remarks=remarks or 'Bulk disbursed.', changed_by=request.user,
            )
            app.status = 'DISBURSED'; app.save(); count += 1
    elif action == 'bulk_export':
        ids_param = '&'.join([f'ids={i}' for i in ids])
        return redirect(f'{request.META.get("HTTP_REFERER", "/reports/export/excel/")}?{ids_param}')
    else:
        messages.error(request, 'Invalid bulk action or insufficient permissions.')
        return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
    messages.success(request, f'{count} application(s) updated via bulk {action.replace("bulk_", "")}.')
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))


# ─── Reports ────────────────────────────────────────────────────

@login_required
def coordinator_reports(request):
    apps = LoanApplication.objects.filter(created_by=request.user)
    status_counts = {s[0]: apps.filter(status=s[0]).count() for s in LoanApplication.Status.choices}
    monthly = (
        apps.annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    context = {
        'total': apps.count(),
        'status_counts': status_counts,
        'status_choices': LoanApplication.Status.choices,
        'monthly_labels': [m['month'].strftime('%b %Y') if m['month'] else '' for m in monthly],
        'monthly_values': [m['count'] for m in monthly],
        'applications': apps.order_by('-created_at'),
    }
    return render(request, 'reports/coordinator.html', context)


@login_required
def reports_dashboard(request):
    if not (request.user.is_superuser or request.user.role in ('GCC_NOIDA', 'SUPER_ADMIN', 'STATE_HEAD')):
        messages.error(request, 'You do not have permission to access reports.')
        return redirect('dashboard')
    base_qs = LoanApplication.objects.all()

    if request.user.role == 'STATE_HEAD' and request.user.state:
        base_qs = base_qs.filter(state=request.user.state)

    state_summary = (
        base_qs.values('state')
        .annotate(
            total=Count('id'),
            approved=Count('id', filter=Q(status='APPROVED')),
            rejected=Count('id', filter=Q(status='REJECTED')),
            disbursed=Count('id', filter=Q(status='DISBURSED')),
            pending=Count('id', filter=Q(status__in=['SUBMITTED', 'STATE_REVIEW', 'RESUBMITTED', 'GCC_REVIEW', 'ADDITIONAL_DOCS_REQUIRED'])),
        )
        .order_by('-total')
    )

    coordinator_perf = (
        base_qs.values('created_by__first_name', 'created_by__last_name', 'created_by__employee_code')
        .annotate(
            total=Count('id'),
            approved=Count('id', filter=Q(status='APPROVED')),
            rejected=Count('id', filter=Q(status='REJECTED')),
        )
        .order_by('-total')
    )

    monthly_volume = (
        base_qs.annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )

    status_split = {
        s[0]: base_qs.filter(status=s[0]).count()
        for s in LoanApplication.Status.choices
    }

    disbursements = base_qs.filter(status='DISBURSED').order_by('-updated_at')[:50]

    context = {
        'state_summary': state_summary,
        'coordinator_perf': coordinator_perf,
        'monthly_labels': [m['month'].strftime('%b %Y') if m['month'] else '' for m in monthly_volume],
        'monthly_values': [m['count'] for m in monthly_volume],
        'status_split': status_split,
        'status_choices': LoanApplication.Status.choices,
        'disbursements': disbursements,
        'status_split_json': json.dumps(status_split),
        'monthly_labels_json': json.dumps([m['month'].strftime('%b %Y') if m['month'] else '' for m in monthly_volume]),
        'monthly_values_json': json.dumps([m['count'] for m in monthly_volume]),
    }
    return render(request, 'reports/dashboard.html', context)


@login_required
def export_excel(request):
    if not (request.user.is_superuser or request.user.role in ('GCC_NOIDA', 'SUPER_ADMIN', 'STATE_HEAD')):
        messages.error(request, 'You do not have permission to export reports.')
        return redirect('dashboard')
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Loan Applications'

    header_font = Font(name='Inter', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    headers = ['App ID', 'Applicant', 'Phone', 'Email', 'PAN', 'Aadhaar',
               'State', 'District', 'Loan Type', 'Amount', 'Tenure (Months)',
               'Status', 'Coordinator', 'Created At', 'Submitted At']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    applications_qs = LoanApplication.objects.all().select_related('created_by')
    if request.user.role == 'STATE_HEAD' and request.user.state:
        applications_qs = applications_qs.filter(state=request.user.state)
    for row_idx, app in enumerate(applications_qs, 2):
        data = [
            app.application_id, app.applicant_name, app.applicant_phone,
            app.applicant_email or '', app.pan_number or '', app.aadhaar_number or '',
            app.state, app.district, app.get_loan_type_display(),
            float(app.loan_amount), app.tenure_months,
            app.get_status_display(),
            app.created_by.get_full_name() if app.created_by else '',
            app.created_at.strftime('%Y-%m-%d %H:%M') if app.created_at else '',
            app.submitted_at.strftime('%Y-%m-%d %H:%M') if app.submitted_at else '',
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = Font(name='Inter', size=10)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="loan_applications.xlsx"'
    wb.save(response)
    return response


@login_required
def export_pdf(request):
    if not (request.user.is_superuser or request.user.role in ('GCC_NOIDA', 'SUPER_ADMIN', 'STATE_HEAD')):
        messages.error(request, 'You do not have permission to export reports.')
        return redirect('dashboard')
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=20)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('Title2', parent=styles['Title'], fontSize=16, spaceAfter=12, textColor=colors.HexColor('#1E3A5F'))
    elements.append(Paragraph('GCC Loan Portal — Application Report', title_style))
    elements.append(Spacer(1, 10))

    header_style = ParagraphStyle('Header', parent=styles['Normal'], fontSize=8, textColor=colors.white, alignment=1)
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=7, alignment=1)

    headers = ['App ID', 'Applicant', 'State', 'District', 'Loan Type', 'Amount', 'Status', 'Coordinator']
    data = [headers]

    pdf_qs = LoanApplication.objects.all().select_related('created_by')
    if request.user.role == 'STATE_HEAD' and request.user.state:
        pdf_qs = pdf_qs.filter(state=request.user.state)
    applications = pdf_qs[:200]
    for app in applications:
        row = [
            app.application_id or '', app.applicant_name or '', app.state or '',
            app.district or '', app.get_loan_type_display() or '',
            f'₹{float(app.loan_amount):,.0f}' if app.loan_amount else '',
            app.get_status_display() or '',
            app.created_by.get_full_name() if app.created_by else '',
        ]
        data.append(row)

    table = Table(data, repeatRows=1, colWidths=[60, 100, 55, 60, 65, 65, 60, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1D4ED8')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f'Generated: {datetime.now().strftime("%d %b %Y, %H:%M")} | Total Records: {len(applications)}',
                              ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7, textColor=colors.grey)))

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="loan_applications.pdf"'
    return response


# ─── Coordinator ────────────────────────────────────────────────

@login_required
def loan_create(request):
    if request.method == 'POST':
        form = LoanApplicationForm(request.POST)
        if form.is_valid():
            app = form.save(commit=False)
            app.created_by = request.user
            app.status = LoanApplication.Status.DRAFT
            app.save()
            ApplicationHistory.objects.create(
                application=app, to_status=LoanApplication.Status.DRAFT,
                changed_by=request.user, remarks='Application created as draft.',
            )
            messages.success(request, 'Loan application draft created successfully.')
            return redirect('loan_document_upload', application_id=app.id)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = LoanApplicationForm()
    return render(request, 'coordinator/loan_form.html', {'form': form})


@login_required
def loan_document_upload(request, application_id):
    application = get_object_or_404(LoanApplication, id=application_id, created_by=request.user)
    documents = application.documents.all()
    if request.method == 'POST':
        form = DocumentUploadForm(request.POST, request.FILES)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.application = application
            doc.uploaded_by = request.user
            doc.original_filename = request.FILES['file'].name
            doc.save()
            messages.success(request, 'Document uploaded successfully.')
            return redirect('loan_document_upload', application_id=application.id)
    else:
        form = DocumentUploadForm()
    return render(request, 'coordinator/document_upload.html', {'application': application, 'form': form, 'documents': documents})


@require_POST
@login_required
def loan_delete_document(request, application_id, doc_id):
    doc = get_object_or_404(DocumentUpload, id=doc_id, application_id=application_id, application__created_by=request.user)
    doc.file.delete()
    doc.delete()
    messages.success(request, 'Document removed.')
    return redirect('loan_document_upload', application_id=application_id)


@login_required
def loan_review_submit(request, application_id):
    application = get_object_or_404(LoanApplication, id=application_id, created_by=request.user)
    documents = application.documents.all()
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'submit':
            if application.status == LoanApplication.Status.DRAFT:
                application.status = LoanApplication.Status.SUBMITTED
                application.submitted_at = __import__('django.utils.timezone', fromlist=['now']).now()
                application.save()
                ApplicationHistory.objects.create(
                    application=application, from_status=LoanApplication.Status.DRAFT,
                    to_status=LoanApplication.Status.SUBMITTED,
                    changed_by=request.user, remarks='Application submitted by Coordinator.',
                )
                messages.success(request, 'Application submitted for review successfully.')
            elif application.status == LoanApplication.Status.RETURNED_FOR_CORRECTION:
                application.status = LoanApplication.Status.RESUBMITTED
                application.save()
                ApplicationHistory.objects.create(
                    application=application, from_status=LoanApplication.Status.RETURNED_FOR_CORRECTION,
                    to_status=LoanApplication.Status.RESUBMITTED,
                    changed_by=request.user, remarks='Application resubmitted after correction.',
                )
                messages.success(request, 'Application resubmitted for review successfully.')
            elif application.status == LoanApplication.Status.ADDITIONAL_DOCS_REQUIRED:
                application.status = LoanApplication.Status.GCC_REVIEW
                application.save()
                ApplicationHistory.objects.create(
                    application=application, from_status=LoanApplication.Status.ADDITIONAL_DOCS_REQUIRED,
                    to_status=LoanApplication.Status.GCC_REVIEW,
                    changed_by=request.user, remarks='Additional documents uploaded, sent back to GCC review.',
                )
                messages.success(request, 'Documents resubmitted for GCC review.')
            return redirect('loan_detail', application_id=application.id)
        elif action == 'save_draft':
            messages.success(request, 'Application saved as draft.')
            return redirect('coordinator_dashboard')
    return render(request, 'coordinator/review_submit.html', {'application': application, 'documents': documents})


def get_activity_log(application):
    events = []
    status_labels = dict(LoanApplication.Status.choices)
    for h in application.history.all():
        if h.from_status and h.to_status and h.from_status != h.to_status:
            events.append({
                'timestamp': h.created_at, 'user': h.changed_by,
                'action': f"Status changed: {status_labels.get(h.from_status, h.from_status)} → {status_labels.get(h.to_status, h.to_status)}",
                'details': h.remarks or '', 'type': 'status_change',
            })
        elif h.remarks:
            events.append({
                'timestamp': h.created_at, 'user': h.changed_by,
                'action': 'Remark added', 'details': h.remarks, 'type': 'remark',
            })
    for doc in application.documents.all():
        events.append({
            'timestamp': doc.uploaded_at, 'user': doc.uploaded_by,
            'action': f"Document uploaded: {doc.get_doc_type_display()}",
            'details': doc.original_filename or '', 'type': 'document_upload',
        })
        if doc.verification_status != 'PENDING' and doc.verified_at:
            events.append({
                'timestamp': doc.verified_at, 'user': doc.verified_by,
                'action': f"Document {doc.get_verification_status_display().lower()}: {doc.get_doc_type_display()}",
                'details': doc.verification_remarks or '', 'type': 'document_verify',
            })
    events.sort(key=lambda e: e['timestamp'], reverse=True)
    return events


@login_required
def loan_detail(request, application_id):
    application = get_object_or_404(LoanApplication, id=application_id)
    if request.user.role == 'COORDINATOR' and application.created_by != request.user:
        messages.error(request, 'You can only view your own applications.')
        return redirect('coordinator_dashboard')
    documents = application.documents.all()
    history = application.history.all()
    remark_form = ApplicationRemarkForm()
    activity_log = get_activity_log(application)

    seen = set()
    status_path = []
    for h in history.order_by('created_at'):
        if h.to_status and h.to_status not in seen:
            status_path.append(h.to_status)
            seen.add(h.to_status)
        if h.from_status and h.from_status not in seen:
            status_path.append(h.from_status)
            seen.add(h.from_status)
    if application.status not in seen:
        status_path.append(application.status)

    status_labels = dict(LoanApplication.Status.choices)
    return render(request, 'coordinator/loan_detail.html', {
        'application': application, 'documents': documents, 'history': history,
        'remark_form': remark_form, 'status_path': status_path,
        'status_labels': status_labels, 'activity_log': activity_log,
    })


@require_POST
@login_required
def loan_add_remark(request, application_id):
    application = get_object_or_404(LoanApplication, id=application_id)
    if request.user.role == 'COORDINATOR' and application.created_by != request.user:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    form = ApplicationRemarkForm(request.POST)
    if form.is_valid():
        ApplicationHistory.objects.create(
            application=application, to_status=application.status,
            remarks=form.cleaned_data['remarks'], changed_by=request.user,
        )
        messages.success(request, 'Remark added.')
    return redirect('loan_detail', application_id=application.id)


# ─── User Management ────────────────────────────────────────────

@login_required
@gcc_or_admin_required
def user_list(request):
    users = User.objects.all().order_by('-date_joined')
    roles = User.Role.choices
    states = [s[0] for s in INDIAN_STATES]
    role_filter = request.GET.get('role', '')
    state_filter = request.GET.get('state', '')
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('q', '')
    if role_filter:
        users = users.filter(role=role_filter)
    if state_filter:
        users = users.filter(state=state_filter)
    if status_filter == 'active':
        users = users.filter(is_active=True)
    elif status_filter == 'inactive':
        users = users.filter(is_active=False)
    if search_query:
        users = users.filter(Q(first_name__icontains=search_query) | Q(last_name__icontains=search_query) | Q(email__icontains=search_query) | Q(employee_code__icontains=search_query))
    role_counts = {
        role: User.objects.filter(role=role).count()
        for role, _ in User.Role.choices
    }
    ul_page = request.GET.get('page', 1)
    ul_paginator = Paginator(users, 20)
    try:
        ul_page_obj = ul_paginator.page(ul_page)
    except (PageNotAnInteger, EmptyPage):
        ul_page_obj = ul_paginator.page(1)
    get_params = request.GET.copy()
    if 'page' in get_params: del get_params['page']
    return render(request, 'users/user_list.html', {
        'ul_page_obj': ul_page_obj, 'roles': roles, 'states': states,
        'role_filter': role_filter, 'state_filter': state_filter,
        'status_filter': status_filter, 'search_query': search_query,
        'role_counts': role_counts,
        'filter_params': get_params.urlencode() if get_params else '',
    })


@login_required
@super_admin_required
def user_create(request):
    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.username = form.cleaned_data['email'] or form.cleaned_data['employee_code']
            user.save()
            AuditLog.objects.create(
                user=request.user, action='CREATE_USER', module='User Management',
                remarks=f'Created user {user.get_full_name()} ({user.role})',
            )
            generated_pwd = getattr(form, 'generated_password', None)
            msg = f'User {user.get_full_name()} created successfully.'
            if generated_pwd:
                msg += f' Generated password: {generated_pwd}'
            messages.success(request, msg)
            return redirect('user_list')
        else:
            errors = []
            for field, errs in form.errors.items():
                errors.extend(errs)
            messages.error(request, ' '.join(errors))
            return redirect('user_list')
    else:
        form = UserForm()
    return render(request, 'users/user_form.html', {'form': form, 'title': 'Add User'})


@login_required
@super_admin_required
def user_edit(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        form = UserForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            AuditLog.objects.create(
                user=request.user, action='EDIT_USER', module='User Management',
                remarks=f'Edited user {user.get_full_name()}',
            )
            messages.success(request, f'User {user.get_full_name()} updated successfully.')
            return redirect('user_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = UserForm(instance=user)
    return render(request, 'users/user_form.html', {'form': form, 'title': 'Edit User', 'edit_user': user})


@login_required
@super_admin_required
@require_POST
def user_toggle_active(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if user == request.user:
        messages.error(request, 'You cannot deactivate your own account.')
        return redirect('user_list')
    user.is_active = not user.is_active
    user.save()
    AuditLog.objects.create(
        user=request.user, action='TOGGLE_USER', module='User Management',
        remarks=f'{"Activated" if user.is_active else "Deactivated"} user {user.get_full_name()}',
    )
    messages.success(request, f'User {user.get_full_name()} {"activated" if user.is_active else "deactivated"} successfully.')
    return redirect('user_list')


@login_required
@super_admin_required
def user_reset_password(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        password = request.POST.get('password')
        if password and len(password) >= 6:
            user.set_password(password)
            user.save()
            AuditLog.objects.create(
                user=request.user, action='RESET_PASSWORD', module='User Management',
                remarks=f'Password reset for user {user.get_full_name()}',
            )
            messages.success(request, f'Password for {user.get_full_name()} reset successfully.')
            return redirect('user_list')
        else:
            messages.error(request, 'Password must be at least 6 characters.')
    return render(request, 'users/reset_password.html', {'reset_user': user})


@login_required
def profile(request):
    from django.contrib.auth.forms import PasswordChangeForm
    form = PasswordChangeForm(request.user)
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            form.save()
            AuditLog.objects.create(
                user=request.user, action='PASSWORD_CHANGE', module='Auth',
                ip_address=request.META.get('REMOTE_ADDR'),
                remarks='Password changed successfully from profile.'
            )
            messages.success(request, 'Your password has been changed successfully.')
            return redirect('profile')
    return render(request, 'profile.html', {'profile_user': request.user, 'password_form': form})


# ─── Audit Trail ─────────────────────────────────────────────────

@login_required
@super_admin_required
def audit_log_view(request):
    logs = AuditLog.objects.all().select_related('user')
    action_filter = request.GET.get('action', '')
    module_filter = request.GET.get('module', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_query = request.GET.get('q', '')

    if action_filter:
        logs = logs.filter(action=action_filter)
    if module_filter:
        logs = logs.filter(module__icontains=module_filter)
    if date_from:
        logs = logs.filter(created_at__gte=date_from)
    if date_to:
        logs = logs.filter(created_at__lte=date_to + ' 23:59:59')
    if search_query:
        logs = logs.filter(
            Q(remarks__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(ip_address__icontains=search_query)
        )

    daily_counts = (
        logs.extra({'date': "date(created_at)"})
        .values('date')
        .annotate(count=Count('id'))
        .order_by('-date')[:30]
    )

    action_breakdown = logs.values('action').annotate(count=Count('id')).order_by('-count')
    module_breakdown = logs.values('module').annotate(count=Count('id')).order_by('-count')

    al_page = request.GET.get('page', 1)
    al_paginator = Paginator(logs.order_by('-created_at'), 25)
    try:
        al_page_obj = al_paginator.page(al_page)
    except (PageNotAnInteger, EmptyPage):
        al_page_obj = al_paginator.page(1)
    get_params = request.GET.copy()
    if 'page' in get_params: del get_params['page']
    context = {
        'al_page_obj': al_page_obj,
        'filter_params': get_params.urlencode() if get_params else '',
        'action_choices': AuditLog.Action.choices,
        'action_filter': action_filter,
        'module_filter': module_filter,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search_query,
        'daily_counts_json': json.dumps(
            [{'date': str(e['date']), 'count': e['count']} for e in daily_counts]
        ),
        'action_breakdown_json': json.dumps(
            [{'action': e['action'], 'count': e['count']} for e in action_breakdown]
        ),
        'module_breakdown_json': json.dumps(
            [{'module': e['module'] or '—', 'count': e['count']} for e in module_breakdown]
        ),
        'total_logs': al_page_obj.paginator.count,
    }
    return render(request, 'audit_log.html', context)


def keep_alive(request):
    if not request.user.is_authenticated:
        from django.http import JsonResponse
        return JsonResponse({'ok': False}, status=403)
    request.session.modified = True
    from django.http import JsonResponse
    return JsonResponse({'ok': True, 'expires': int(request.session.get_expiry_date().timestamp())})
